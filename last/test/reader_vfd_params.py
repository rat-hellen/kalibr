import serial
import crc_vfd_ascii as crc_vfd_ascii
from openpyxl import Workbook
import yaml
import time

#with open("lotok.yaml", 'r') as stream:
#    data_loaded = yaml.safe_load(stream)
#    print(data_loaded)
#time.sleep(1000)

excel_file = Workbook()
excel_sheet = excel_file.create_sheet(title='VFD', index=0)

serial_port_vfd = None
port = '/dev/ttyMXUSB4'
try:
    serial_port_vfd = serial.Serial(
    port=port,
    baudrate=9600,
    bytesize=serial.EIGHTBITS,
    parity=serial.PARITY_NONE,
    stopbits=serial.STOPBITS_TWO,
    timeout=1, 
    write_timeout=0,
    xonxoff=False,
    rtscts=False,
    dsrdtr=False)
except:
    print('Serial_port_VFD', port, 'No Open')
else:
    print('Serial_port_VFD', port, 'Open')

def SendCommand03(addr,data):
    buffer = b':0103'+ addr + data
    cs = crc_vfd_ascii.lrc(buffer)
    buffer+= cs + b'\r\n'
    #print(buffer)

    serial_port_vfd.write(buffer)
    d = serial_port_vfd.read(17)
    #print('ret:',d)
    return d

count_param =  {0:15, 1:24, 2:14, 3:16, 4:25, 5:33, 6:18, 7:10, 8:21, 9:9, 10:12, 11:12}
row=0
for i in range(len(count_param)):
    a = "%02d" %i
    for j in range(count_param[i]):
        b = "%02d" %j
        addr = bytes(a, encoding="ascii")+bytes(b, encoding="ascii")
        par = SendCommand03(addr,b'0001')
        par_to_save = par.replace(b'\r\n',b'')
        par_hex = ''
        par_dec = ''
        par_bin = ''
        dec = 0
        if par.find(b':010302') != -1:
            #print(addr, par, par_to_save[7:-2])
            par_hex = par_to_save[7:-2]
            dec = int(par_hex,16)
            par_dec = str(dec)
            par_bin = bin(dec)
            par_bin = par_bin.replace('0b','')
            #par_bin = "{0:b}".format(dec)
        else:
            par_hex = '#'
            par_bin = '#'
        row+=1
        print(row, addr, par_hex, par_dec, par_bin)
        excel_sheet.cell(row=row, column=1).value = addr
        excel_sheet.cell(row=row, column=2).value = par_to_save
        excel_sheet.cell(row=row, column=3).value = par_hex
        excel_sheet.cell(row=row, column=4).value = par_dec
        excel_sheet.cell(row=row, column=5).value = par_bin
excel_file.save(filename="vfd/vfd_params.xlsx")


